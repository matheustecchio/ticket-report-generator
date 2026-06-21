from __future__ import annotations

import re
import shutil
import subprocess
from copy import copy
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ModuleNotFoundError as exc:
    raise SystemExit("Missing dependency: install openpyxl with 'pip install openpyxl'.") from exc


BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_PATH = BASE_DIR / "report-template.xlsx"
OUTPUT_DIR = BASE_DIR / "temp"


def print_header() -> None:
    print("=" * 36)
    print(" Ticket Report Generator")
    print("=" * 36)


def ask_required(prompt: str) -> str:
    while True:
        value = input(prompt).strip()
        if value:
            return value
        print("This field is required.")


def ask_date(prompt: str, required: bool = True) -> str:
    while True:
        value = input(prompt).strip()
        if not value and not required:
            return ""

        try:
            datetime.strptime(value, "%d/%m/%Y")
            return value
        except ValueError:
            print("Invalid date. Use dd/mm/yyyy.")


def ask_yes_no(prompt: str, default: bool = False) -> bool:
    while True:
        value = input(prompt).strip().lower()
        if not value:
            return default
        if value in {"y", "yes"}:
            return True
        if value in {"n", "no"}:
            return False
        print("Please answer y or n.")


def collect_ticket_info() -> dict[str, str]:
    return {
        "ticket_id": ask_required("Ticket ID: "),
        "reporter": ask_required("Reporter: "),
        "title": ask_required("Title: "),
        "date": ask_date("Date (dd/mm/yyyy): "),
        "assignee": ask_required("Assignee: "),
        "state": ask_required("State: "),
        "time_spent": ask_required("Time Spent: "),
        "solved_on": ask_date("Solved On (dd/mm/yyyy): ", required=False),
        "due_date": ask_date("Due Date (dd/mm/yyyy): "),
    }


def collect_issue() -> dict[str, str]:
    return {
        "issue": ask_required("Issue: "),
        "action": ask_required("Action: "),
        "status": ask_required("Status: "),
    }


def collect_issues() -> list[dict[str, str]]:
    issues = [collect_issue()]

    while ask_yes_no("Add another issue [y/N]: "):
        issues.append(collect_issue())

    return issues


def fill_workbook(ticket_info: dict[str, str], issues: list[dict[str, str]]) -> Path:
    workbook = load_workbook(TEMPLATE_PATH)
    sheet = workbook.active

    sheet["A11"] = ticket_info["ticket_id"]
    sheet["B11"] = ticket_info["reporter"]
    sheet["D11"] = ticket_info["title"]
    sheet["H11"] = ticket_info["date"]
    sheet["A14"] = ticket_info["assignee"]
    sheet["D14"] = ticket_info["state"]
    sheet["F14"] = ticket_info["time_spent"]
    sheet["G14"] = ticket_info["solved_on"]
    sheet["H14"] = ticket_info["due_date"]

    for index, issue in enumerate(issues, start=18):
        if index > 18:
            sheet.insert_rows(index)
            copy_summary_row_style(sheet, 18, index)
            copy_summary_row_merges(sheet, 18, index)

        sheet[f"A{index}"] = issue["issue"]
        sheet[f"D{index}"] = issue["action"]
        sheet[f"H{index}"] = issue["status"]

    OUTPUT_DIR.mkdir(exist_ok=True)
    xlsx_path = OUTPUT_DIR / f"{safe_filename(ticket_info['ticket_id'])}.xlsx"
    workbook.save(xlsx_path)
    return xlsx_path


def copy_summary_row_style(sheet, source_row: int, target_row: int) -> None:
    for column in range(1, sheet.max_column + 1):
        source_cell = sheet.cell(source_row, column)
        target_cell = sheet.cell(target_row, column)

        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.font:
            target_cell.font = copy(source_cell.font)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        if source_cell.border:
            target_cell.border = copy(source_cell.border)

    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height


def copy_summary_row_merges(sheet, source_row: int, target_row: int) -> None:
    for merged_range in list(sheet.merged_cells.ranges):
        if merged_range.min_row == source_row and merged_range.max_row == source_row:
            sheet.merge_cells(
                start_row=target_row,
                start_column=merged_range.min_col,
                end_row=target_row,
                end_column=merged_range.max_col,
            )


def safe_filename(value: str) -> str:
    filename = re.sub(r"[^A-Za-z0-9._-]+", "-", value.strip())
    return filename.strip(".-_") or "ticket-report"


def convert_to_pdf(xlsx_path: Path) -> Path:
    converter = shutil.which("libreoffice") or shutil.which("soffice")
    if not converter:
        raise RuntimeError("LibreOffice is required to generate PDFs.")

    result = subprocess.run(
        [
            converter,
            "--headless",
            "--convert-to",
            "pdf",
            "--outdir",
            str(OUTPUT_DIR),
            str(xlsx_path),
        ],
        check=False,
        capture_output=True,
        text=True,
    )
    if result.returncode != 0:
        message = result.stderr.strip() or result.stdout.strip()
        raise RuntimeError(f"Failed to generate PDF: {message}")

    pdf_path = xlsx_path.with_suffix(".pdf")
    if not pdf_path.exists():
        raise RuntimeError("PDF conversion finished, but the PDF file was not created.")

    return pdf_path


def create_report() -> Path:
    ticket_info = collect_ticket_info()
    print()
    issues = collect_issues()

    xlsx_path = fill_workbook(ticket_info, issues)
    try:
        pdf_path = convert_to_pdf(xlsx_path)
    finally:
        xlsx_path.unlink(missing_ok=True)

    print(f"Ticket Report created at {pdf_path}")
    return pdf_path


def main() -> None:
    print_header()
    print()

    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Template not found: {TEMPLATE_PATH}")

    while True:
        create_report()
        print()
        if not ask_yes_no("Create another ticket report [y/N]: "):
            break
        print()


if __name__ == "__main__":
    main()
